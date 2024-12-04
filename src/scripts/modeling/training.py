import xgboost as xgb
import pandas as pd
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.ticker import AutoMinorLocator
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.metrics import confusion_matrix, accuracy_score, balanced_accuracy_score, precision_score, recall_score, f1_score, classification_report

from misc import utils
from misc.utils import BaseClass, get_alpaca_secrets, load_config, log, read_and_duplicate, read_df, save_df_as_parquet, save_df_as_csv, get_all_paths_from_loc, get_name_and_type
from misc.features_calc import update_config, load_features_config, get_paths_and_cols_from_config, check_index_is_monotonic_increasing
# from misc.features_calc import 

import os
import re
import boto3
import warnings
import datetime
import numpy as np
import pandas as pd
from tqdm import tqdm
from sklearn.linear_model import LinearRegression
import matplotlib.pyplot as plt
from matplotlib.ticker import (AutoMinorLocator, MultipleLocator)
warnings.filterwarnings("ignore")





class ModelTraining(BaseClass):
    def __init__(self, config=None):
        super().__init__()
        self.name = 'ModelTrainingFilePrep'
        log(f'running {self.name}')
        if config is None:
            self.config = load_config()
        else:
            self.config = config

        self.paths_config = self.config['paths_config']

        self.bucket_loc = f's3://{self.paths_config["s3_bucket"]}'
        self.base_folder = self.paths_config["base_folder"]
        self.data_folder = self.paths_config["data_folder"]
        self.data_prep_folder = self.paths_config["data_prep_folder"]
        self.reduced_autocorelation_folder = self.paths_config["reduced_autocorelation_folder"]
        self.feature_prep_folder = self.paths_config["feature_prep_folder"]
        self.client = boto3.client('s3')

        self.common_config = self.config['technical_yaml']['common']
        self.feature_prep = self.config['technical_yaml']['feature_prep']
        self.modeling_config = self.config['technical_yaml']['modeling']

        self.features_to_be_calculated = self.feature_prep['features_to_be_calculated']
        self.symbols = self.feature_prep['symbols']
        self.calc_metrics = self.feature_prep['calc_metrics']
        
        self.features_listing_config = load_features_config()
        
    def extract(self):
        
        model = xgb.XGBClassifier(n_estimators=100,
                                objective='multi:softmax',
                                n_jobs =-1,
                                random_state=420,
                                num_class=3,
                                eval_metric=['merror','mlogloss'])
        
        df = read_df('s3://sisyphus-general-bucket/AthenaInsights/latest_data/model/data/stock_bars_1min_diff.parquet' )
        
        df = df.fillna(0)
        category_map = self.modeling_config['category_map']
        reverse_category_map = {v: k for k, v in category_map.items()}
        df['mapped_category'] = df['category'].map({'A': 0, 'B': 1, 'C':2})
        df['mapped_category'].value_counts()
        
        start_date = self.modeling_config['start_date']
        end_date = self.modeling_config['send_date']
        date_series = pd.date_range(start=start_date, end=end_date, freq='D')
        date_series = [z.strftime('%Y-%m-%d') for z in date_series]
        
        log(f'start_date - {start_date}')
        log(f'end_date - {end_date}')
        log(f'dates - {date_series}')
        
        return {
            'df': df,
            'model': model,
            'date_series': date_series,
        }

    @staticmethod
    def get_dates(dt):
        test_date = dt
        next_day = (datetime.datetime.strptime(test_date, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')
        next_10_day = (datetime.datetime.strptime(test_date, '%Y-%m-%d') + timedelta(days=9)).strftime('%Y-%m-%d')
        prev_day = (datetime.datetime.strptime(test_date, '%Y-%m-%d') + timedelta(days=-1)).strftime('%Y-%m-%d')
        return test_date, next_day, next_10_day, prev_day
    
    @staticmethod
    def get_train_test_split(df, test_date, next_day, next_10_day, prev_day ):
        X_train = df.loc[:prev_day, ].drop(columns=['category', 'mapped_category'])
        y_train = df.loc[:prev_day, 'mapped_category']

        X_test_only_next_day = df.loc[test_date, ].drop(columns=['category', 'mapped_category'])
        y_test_only_next_day = df.loc[test_date, 'mapped_category']

        X_test_next_10_days = df.loc[test_date: next_10_day, ].drop(columns=['category', 'mapped_category'])
        y_test_next_10_days = df.loc[test_date: next_10_day, 'mapped_category']

        X_test_full = df.loc[test_date:, ].drop(columns=['category', 'mapped_category'])
        y_test_full = df.loc[test_date:, 'mapped_category']

        return X_train, y_train, X_test_only_next_day, y_test_only_next_day, X_test_next_10_days, y_test_next_10_days, X_test_full, y_test_full
    
    @staticmethod
    def initialte_and_train(model, X_train, y_train, X_test_only_next_day, y_test_only_next_day, 
                            X_test_next_10_days, y_test_next_10_days, X_test_full, y_test_full):
        model.fit(X_train,
                y_train,
                verbose=0,
                eval_set=[(X_train, y_train), (X_test_only_next_day, y_test_only_next_day), (X_test_next_10_days, y_test_next_10_days), (X_test_full, y_test_full)])

        return model
    
    def train(self, dfs):
        df = dfs['df']
        model = dfs['model']
        date_series = ['date_series']
        
        for dt in tqdm(date_series):
            # print(f"running for dt = {dt}", dt)
            log(f"running for dt = {dt}")
            test_date, next_day, next_10_day, prev_day = self.get_dates(dt)
            log(f"test_date: {test_date}, next_day: {next_day}, next_10_day: {next_10_day}, prev_day: {prev_day}")

            X_train, y_train, X_test_only_next_day, y_test_only_next_day, X_test_next_10_days, y_test_next_10_days, X_test_full, y_test_full = get_train_test_split(df, test_date, next_day, next_10_day, prev_day)
            if y_test_full.empty:
                break

            log(f"y_train.value_counts():{y_train.value_counts()}")
            log(f"y_test_only_next_day.value_counts():\n{y_test_only_next_day.value_counts()}")
            log(f"y_test_next_10_days.value_counts():\n{y_test_next_10_days.value_counts()}")
            log(f"y_test_full.value_counts():\n{y_test_full.value_counts()}")

            log("training the model")
            clf = initialte_and_train(mocdl, X_train, y_train, X_test_only_next_day, y_test_only_next_day, X_test_next_10_days, y_test_next_10_days, X_test_full, y_test_full)
            log(f"model trained for {dt}")

            log("results")
            get_results(clf, dt)

            if not X_test_only_next_day.empty:
                log("1 day test")
                generate_reports(X_test_only_next_day, y_test_only_next_day, clf, dur='1_day', dt)

            if not X_test_next_10_days.empty:
                log("10 day test")
                generate_reports(X_test_next_10_days, y_test_next_10_days, clf, dur='10_day', dt)

            if not X_test_full.empty:
                log("full test")
                generate_reports(X_test_full, y_test_full, clf, next_day=True, dur='full_day', dt)

            feature_important = clf.feature_importances_ 
            keys = list(X_train.columns)
            values = list(feature_important)

            log("feature importances")
            fea_imp = pd.DataFrame(data=values, index=keys, columns=["score"]).sort_values(by = "score", ascending=True)
            log(fea_imp, dt)

            if not X_test_only_next_day.empty:
                preds_probs = clf.predict_proba(X_test_only_next_day)
                preds_probs1 = (preds_probs >= 0.5).argmax(axis=1,)
                plot_categorization(df, dt, pd.Series(preds_probs1).map(reverse_category_map))
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
    def save(self, dfs):
        results_loc = 'results'
        model_folder = 'model_2_diff_prices_without_rsi'
        directory = f"{results_loc}/{model_folder}"
        images_directory = f"{directory}/images"

        if not os.path.exists(directory):
            os.mkdir(directory)

        if not os.path.exists(images_directory):
            os.mkdir(images_directory)

        file_name = 'results_0.xlsx'
        file_name = f"{directory}/results_0.xlsx"

        # if os.path.exists(file_name):
        #     file_name, result_num = file_name.split(".")[0].split("_")
        #     result_num += 1
        #     file_name = file_name + result_num + ".xlsx"
        #     print(f'base name: {file_name}')


        data = {'features used': list(df.columns)}
        data = pd.DataFrame(data)

        sheet_name = 'InitialSheet'

        # Write the DataFrame to an Excel file with a custom sheet name
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            data.to_excel(writer, sheet_name=sheet_name, index=False)
    

        
        
        
        
        
        




def log(x, dt, image=None, add=None):
    if type(x)==type('str') or type(x).__name__=='DataFrame':
        if type(x)==type('str'):
            data = {'text': [x]}
            data = pd.DataFrame(data)
        else:
            data = x

        book = openpyxl.load_workbook(file_name)
        if dt in book.sheetnames:
            sheet = book[dt]
            start_row = sheet.max_row + 1  # Find the first empty row
            
        else:
            sheet = book.create_sheet(dt)  # Create a new sheet
            start_row = 1

         # Convert DataFrame to rows and append to the sheet
        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        book.save(file_name)

    elif image==1:
        book = openpyxl.load_workbook(file_name)
        sheet = book[dt]
        img = Image(add)
        sheet.add_image(img)
        book.save(file_name)
    else:
        raise ValueError('what are you trying to save?')
        



def get_results(clf, dt):
    results = clf.evals_result()
    epochs = len(results['validation_0']['mlogloss'])
    x_axis = range(0, epochs)

    log(f"results: {results}", dt)
    log(f"epochs: {epochs}", dt)
    log("\n\n", dt)

    # xgboost 'mlogloss' plot
    fig, ax = plt.subplots(figsize=(9,5))
    ax.plot(x_axis, results['validation_0']['mlogloss'], label='Train')
    ax.plot(x_axis, results['validation_1']['mlogloss'], label='Test')
    ax.plot(x_axis, results['validation_2']['mlogloss'], label='Test_full')
    ax.plot(x_axis, results['validation_3']['mlogloss'], label='Test_only_next_day')
    ax.legend()
    plt.ylabel('mlogloss')
    plt.title(f'GridSearchCV XGBoost mlogloss - {dt}')
    # plt.show()
    fig.savefig(f'{images_directory}/GridSearchCV XGBoost mlogloss - {dt}.png')
    # log(None, dt, image=1, add=f'results/images/GridSearchCV XGBoost mlogloss - {dt}.png')
    log("\n\n", dt)

    # xgboost 'merror' plot
    fig, ax = plt.subplots(figsize=(9,5))
    ax.plot(x_axis, results['validation_0']['merror'], label='Train')
    ax.plot(x_axis, results['validation_1']['merror'], label='Test')
    ax.plot(x_axis, results['validation_2']['merror'], label='Test_full')
    ax.plot(x_axis, results['validation_3']['merror'], label='Test_only_next_day')
    ax.legend()
    plt.ylabel('merror')
    plt.title(f'GridSearchCV XGBoost merror - {dt}')
    # plt.show()
    fig.savefig(f'{images_directory}/GridSearchCV XGBoost merror - {dt}.png')
    # log(None, dt, image=1, add=f'results/images/GridSearchCV XGBoost merror - {dt}.png')
    log("\n\n", dt)
    
    
def generate_reports(x, y, clf, next_day=False):
    log('## ---------- Model Classification Report ----------', dt)
    log('## get predictions and create model quality report', dt)

    y_p = clf.predict(x)

    log('\n------------------ Confusion Matrix -----------------\n', dt)
    log(pd.DataFrame(confusion_matrix(y, y_p)), dt)

    if next_day:
        preds_probs = clf.predict_proba(x)
        for i in range(4, 10, 1):
            log(f'threshold - {i/10}', dt)
            preds_probs1 = (preds_probs>=i/10).argmax(axis=1,)
            log(pd.DataFrame(confusion_matrix(y, preds_probs1)), dt)
            log('\n\n', dt)

    log('\n-------------------- Key Metrics --------------------', dt)
    log('\nAccuracy: {:.2f}'.format(accuracy_score(y, y_p)), dt)
    log('Balanced Accuracy: {:.2f}\n'.format(balanced_accuracy_score(y, y_p)), dt)

    log('Micro Precision: {:.2f}'.format(precision_score(y, y_p, average='micro')), dt)
    log('Micro Recall: {:.2f}'.format(recall_score(y, y_p, average='micro')), dt)
    log('Micro F1-score: {:.2f}\n'.format(f1_score(y, y_p, average='micro')), dt)

    log('Macro Precision: {:.2f}'.format(precision_score(y, y_p, average='macro')), dt)
    log('Macro Recall: {:.2f}'.format(recall_score(y, y_p, average='macro')), dt)
    log('Macro F1-score: {:.2f}\n'.format(f1_score(y, y_p, average='macro')), dt)

    log('Weighted Precision: {:.2f}'.format(precision_score(y, y_p, average='weighted')), dt)
    log('Weighted Recall: {:.2f}'.format(recall_score(y, y_p, average='weighted')), dt)
    log('Weighted F1-score: {:.2f}'.format(f1_score(y, y_p, average='weighted')), dt)

    log('\n--------------- Classification Report ---------------\n', dt)
    log(classification_report(y, y_p), dt)
    log('---------------------- XGBoost ----------------------', dt) # unnecessary fancy styling
    
def plot_categorization(df, date, pred, field='close', ):
    """ Plot categorization for a given day with dynamic field selection """
    df_day = df.loc[date]
    df_day['preds'] = list(pred)

    plt.figure(figsize=(14, 7))
    fig, axs = plt.subplots(2, 1, figsize=(14,14))
    axs[0].plot(df_day.index, df_day[field], label=f'{field.capitalize()} Price', color='gray', linewidth=2)
    for cat, color in zip(['A', 'B', 'C'], ['green', 'red', 'gray']):
        axs[0].scatter(df_day[df_day['category'] == cat].index,
                       df_day[df_day['category'] == cat][field],
                       color=color, label=f'Category {cat}',
                       s=30 if cat != 'C' else 0)
    axs[0].grid(axis='x', which='major', linestyle=':', linewidth='0.5', color='gray')
    axs[0].grid(axis='x', which='minor', linestyle=':', linewidth='0.5', color='gray')
    axs[0].xaxis.set_minor_locator(AutoMinorLocator(n=10))

    axs[1].plot(df_day.index, df_day[field], label=f'{field.capitalize()} Price', color='gray', linewidth=2)
    for cat, color in zip(['A', 'B', 'C'], ['green', 'red', 'gray']):
        axs[1].scatter(df_day[df_day['preds'] == cat].index,
                       df_day[df_day['preds'] == cat][field],
                       color=color, label=f'Preds {cat}',
                       s=20 if cat != 'C' else 0)
    axs[1].grid(axis='x', which='major', linestyle=':', linewidth='0.5', color='gray')
    axs[1].grid(axis='x', which='minor', linestyle=':', linewidth='0.5', color='gray')
    axs[1].xaxis.set_minor_locator(AutoMinorLocator(n=10))

    plt.legend()
    plt.title(f'Price Categorization on {date}')
    plt.xlabel('Timestamp')
    plt.ylabel(f'{field.capitalize()} Price')
    # plt.show()
    plt.savefig(f'{images_directory}/plot for day - {date}.png')
    # log(None, dt, image=1, add=f'results/images/plot for day - {date}.png')
    log("\n\n", date)